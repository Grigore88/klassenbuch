"""
Klassenbuch – PostgreSQL Data Extractor + Berichtsheft Excel Export
Uses the original template as a base for pixel-perfect formatting.
"""

from __future__ import annotations
import os
import shutil
import getpass
from dataclasses import dataclass, field
from datetime import date, timedelta
from collections import defaultdict
from typing import Optional

try:
    import tomllib          # Python 3.11+
except ModuleNotFoundError:
    try:
        import tomli as tomllib  # pip install tomli
    except ModuleNotFoundError:
        tomllib = None       # falls back gracefully

import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook

def _find_template() -> str:
    """Locate berichtsheft_template.xlsx next to this script (works on Windows too)."""
    candidates = [
        # Same folder as the .py file
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "berichtsheft_template.xlsx"),
        # Current working directory (useful when running from a different folder)
        os.path.join(os.getcwd(), "berichtsheft_template.xlsx"),
    ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    # Not found – return first candidate so the error message is helpful
    return candidates[0]

TEMPLATE_PATH = _find_template()

# ─────────────────────────────────────────────────────────
# Data Classes
# ─────────────────────────────────────────────────────────

@dataclass
class Unterrichtseinheit:
    einheit_id: int
    stunde: int
    inhalt: str

    def __str__(self):
        return f"  Stunde {self.stunde}: {self.inhalt or '(kein Inhalt)'}"


@dataclass
class Dozent:
    dozent_id: int
    vorname: str
    nachname: str

    @property
    def full_name(self):
        return f"{self.vorname} {self.nachname}"

    def __str__(self):
        return self.full_name


@dataclass
class Lernfeld:
    lernfeld_id: str
    titel: str
    start_datum: Optional[date]
    end_datum: Optional[date]
    dozenten: list[Dozent] = field(default_factory=list)

    def __str__(self):
        dozs = ", ".join(d.full_name for d in self.dozenten) or "keine"
        return f"{self.lernfeld_id}: {self.titel} ({self.start_datum} -> {self.end_datum}) | Dozenten: {dozs}"


@dataclass
class Lerntag:
    lerntag_id: int
    datum: date
    lernfeld: Optional[Lernfeld]
    dozent: Optional[Dozent]
    einheiten: list[Unterrichtseinheit] = field(default_factory=list)

    @property
    def calendar_week(self):
        return self.datum.isocalendar()[1]

    @property
    def year(self):
        return self.datum.isocalendar()[0]

    @property
    def year_week(self):
        return f"{self.year}-KW{self.calendar_week:02d}"

    def __str__(self):
        lf  = self.lernfeld.lernfeld_id if self.lernfeld else "—"
        doz = self.dozent.full_name     if self.dozent   else "—"
        lines = [f"📅 {self.datum}  [{self.year_week}]  LF: {lf}  Dozent: {doz}"]
        for e in sorted(self.einheiten, key=lambda x: x.stunde):
            lines.append(str(e))
        return "\n".join(lines)


# ─────────────────────────────────────────────────────────
# Database – config loader
# Precedence (highest → lowest):
#   CLI / explicit kwarg  >  config.toml  >  PG* environment variables
# ─────────────────────────────────────────────────────────

def _find_config() -> str | None:
    """Look for config.toml next to the script, then in the working directory."""
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.toml"),
        os.path.join(os.getcwd(), "config.toml"),
    ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    return None


def _load_config(config_path: str | None = None) -> dict:
    """
    Load [database] section from config.toml.
    Returns an empty dict if tomllib is unavailable or file not found.
    """
    if tomllib is None:
        return {}
    path = config_path or _find_config()
    if not path:
        return {}
    try:
        with open(path, "rb") as f:
            data = tomllib.load(f)
        return data.get("database", {})
    except Exception as e:
        print(f"⚠️  Could not read config.toml: {e}")
        return {}


def build_db_config(config_path: str | None = None, **overrides) -> dict:
    """
    Assemble the psycopg2 connection dict using the three-tier precedence:
      overrides  >  config.toml [database]  >  PG* environment variables

    Parameters
    ----------
    config_path : explicit path to config.toml  (optional)
    **overrides : any key from the dict below, e.g. host='myserver'
    """
    toml_cfg = _load_config(config_path)

    cfg = {
        "host":     overrides.get("host")
                    or toml_cfg.get("host")
                    or os.getenv("PGHOST", "localhost"),
        "port":     int(overrides.get("port")
                    or toml_cfg.get("port")
                    or os.getenv("PGPORT", 5432)),
        "dbname":   overrides.get("dbname")
                    or toml_cfg.get("name")       # config.toml uses "name"
                    or os.getenv("PGDATABASE", "klassenbuch"),
        "user":     overrides.get("user")
                    or toml_cfg.get("user")
                    or os.getenv("PGUSER", "postgres"),
        "password": overrides.get("password")
                    or toml_cfg.get("password")
                    or os.getenv("PGPASSWORD", ""),
    }

    # If password is still empty, prompt interactively (never store blank credentials)
    if not cfg["password"]:
        try:
            cfg["password"] = getpass.getpass(
                f"Password for {cfg['user']}@{cfg['host']}/{cfg['dbname']}: "
            )
        except (EOFError, KeyboardInterrupt):
            cfg["password"] = ""   # non-interactive environment, leave blank

    return cfg


# Module-level default built once at import time
DB_CONFIG = build_db_config()


def get_connection(db_config: dict | None = None):
    return psycopg2.connect(**(db_config or DB_CONFIG))


def fetch_all() -> list[Lerntag]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("SELECT * FROM dozent ORDER BY dozent_id")
            dozent_map: dict[int, Dozent] = {
                r["dozent_id"]: Dozent(r["dozent_id"], r["vorname"], r["nachname"])
                for r in cur.fetchall()
            }

            cur.execute("SELECT * FROM lernfeld ORDER BY lernfeld_id")
            lernfeld_map: dict[str, Lernfeld] = {
                r["lernfeld_id"]: Lernfeld(
                    r["lernfeld_id"], r["titel"], r["start_datum"], r["end_datum"]
                )
                for r in cur.fetchall()
            }

            cur.execute("SELECT * FROM lernfeld_dozent")
            for r in cur.fetchall():
                lf  = lernfeld_map.get(r["lernfeld_id"])
                doz = dozent_map.get(r["dozent_id"])
                if lf and doz:
                    lf.dozenten.append(doz)

            cur.execute(
                "SELECT lerntag_id, datum, lernfeld_id, dozent_id FROM lerntag ORDER BY datum"
            )
            lerntag_map: dict[int, Lerntag] = {}
            for r in cur.fetchall():
                lerntag_map[r["lerntag_id"]] = Lerntag(
                    lerntag_id=r["lerntag_id"],
                    datum=r["datum"],
                    lernfeld=lernfeld_map.get(r["lernfeld_id"]),
                    dozent=dozent_map.get(r["dozent_id"]),
                )

            cur.execute(
                "SELECT einheit_id, lerntag_id, stunde, inhalt "
                "FROM unterrichtseinheit ORDER BY lerntag_id, stunde"
            )
            for r in cur.fetchall():
                lt = lerntag_map.get(r["lerntag_id"])
                if lt:
                    lt.einheiten.append(
                        Unterrichtseinheit(r["einheit_id"], r["stunde"], r["inhalt"] or "")
                    )

    return list(lerntag_map.values())


# ─────────────────────────────────────────────────────────
# Grouping helpers
# ─────────────────────────────────────────────────────────

def group_by_lernfeld(lerntage: list[Lerntag]) -> dict[str, list[Lerntag]]:
    groups: dict[str, list[Lerntag]] = defaultdict(list)
    for lt in lerntage:
        key = lt.lernfeld.lernfeld_id if lt.lernfeld else "—"
        groups[key].append(lt)
    return dict(sorted(groups.items()))


def group_by_calendar_week(lerntage: list[Lerntag]) -> dict[str, list[Lerntag]]:
    groups: dict[str, list[Lerntag]] = defaultdict(list)
    for lt in lerntage:
        groups[lt.year_week].append(lt)
    return dict(sorted(groups.items()))


# ─────────────────────────────────────────────────────────
# Excel Berichtsheft Export  (template-based)
# ─────────────────────────────────────────────────────────

# Template layout – first Excel row for each weekday block (Mon=0 … Fri=4)
_DAY_FIRST_ROW = {0: 4, 1: 15, 2: 26, 3: 37, 4: 48}
_ROWS_PER_DAY  = 11
_CONTENT_COL   = 2   # column B
_HOURS_COL     = 10  # column J

_ROW_HEIGHT_PT = 12.5   # matches template default row height


def _clear_data_cells(ws) -> None:
    """Erase variable content cells, leaving all formatting intact."""
    ws.cell(row=1, column=4).value = None   # Nr. + number  (D1)
    ws.cell(row=1, column=6).value = None   # date + LF info (F1)
    ws.cell(row=2, column=6).value = None   # Ausbilder + name (F2)

    for first_row in _DAY_FIRST_ROW.values():
        for offset in range(_ROWS_PER_DAY):
            row = first_row + offset
            ws.cell(row=row, column=_CONTENT_COL).value = None
            ws.cell(row=row, column=_HOURS_COL).value   = None
            ws.row_dimensions[row].height = None  # let Excel auto-fit

    ws.cell(row=59, column=_HOURS_COL).value = None


def _fill_sheet(ws, week_lerntage, kw_number, kw_year, report_nr) -> None:
    """Write week data into the cleared template sheet."""

    day_map: dict[int, Lerntag] = {}
    for lt in week_lerntage:
        wd = lt.datum.weekday()
        if 0 <= wd <= 4:
            day_map[wd] = lt

    lf_sample  = next((lt.lernfeld for lt in week_lerntage if lt.lernfeld), None)
    doz_sample = next((lt.dozent   for lt in week_lerntage if lt.dozent),   None)

    # ── Actual Mon–Fri dates of this calendar week ────────
    jan4     = date(kw_year, 1, 4)
    week_mon = jan4 + timedelta(
        weeks=kw_number - jan4.isocalendar()[1],
        days=-jan4.weekday()
    )
    week_fri = week_mon + timedelta(days=4)
    date_range = f"{week_mon.strftime('%d.%m.')} \u2013 {week_fri.strftime('%d.%m.%Y')}"

    # H1 (merged H:J): date range on line 1, LF title on line 2
    lf_part  = f"\n{lf_sample.lernfeld_id} – {lf_sample.titel}" if lf_sample else ""
    h1_value = date_range + lf_part

    # ── Fill header ───────────────────────────────────────
    ws.cell(row=1, column=4).value = f"Nr. {report_nr}"          # D1:E1
    ws.cell(row=1, column=6).value = h1_value                    # F1:J1
    doz_name = doz_sample.full_name if doz_sample else ""
    ws.cell(row=2, column=6).value = f"Ausbilder: {doz_name}"   # F2:J2

    # ── Fill day blocks ───────────────────────────────────
    total_hours = 0
    for day_idx, first_row in _DAY_FIRST_ROW.items():
        lt        = day_map.get(day_idx)
        einheiten = sorted(lt.einheiten, key=lambda e: e.stunde) if lt else []

        # Warn if data would be silently truncated
        if len(einheiten) > _ROWS_PER_DAY:
            day_name = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag"][day_idx]
            print(f"  ⚠️  {day_name} KW{kw_number}: {len(einheiten)} Einheiten, "
                  f"nur die ersten {_ROWS_PER_DAY} werden exportiert.")

        for offset in range(_ROWS_PER_DAY):
            row = first_row + offset
            if offset < len(einheiten):
                inhalt = einheiten[offset].inhalt
                ws.cell(row=row, column=_CONTENT_COL).value = inhalt
                ws.cell(row=row, column=_HOURS_COL).value   = 1
                total_hours += 1

    ws.cell(row=59, column=_HOURS_COL).value = total_hours
    ws.title = f"KW{kw_number:02d} {kw_year}"


def create_berichtsheft(
    lerntage: list[Lerntag],
    year_week: str,
    output_path: str,
    report_nr: int = 1,
    template_path: str = TEMPLATE_PATH,
) -> None:
    """
    Generate a Berichtsheft .xlsx for one calendar week.

    Parameters
    ----------
    lerntage      : full list of Lerntag objects
    year_week     : '2025-KW04'
    output_path   : destination .xlsx path
    report_nr     : sequential number printed in header cell E1
    template_path : path to berichtsheft_template.xlsx
    """
    by_kw     = group_by_calendar_week(lerntage)
    week_days = by_kw.get(year_week, [])

    kw_year_str, kw_str = year_week.split("-KW")
    kw_number = int(kw_str)
    kw_year   = int(kw_year_str)

    if not os.path.isfile(template_path):
        raise FileNotFoundError(
            f"Template not found:\n  {template_path}\n\n"
            "Make sure 'berichtsheft_template.xlsx' is in the same folder as this script.\n"
            "Or pass template_path='C:/full/path/berichtsheft_template.xlsx' explicitly."
        )
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    _clear_data_cells(ws)
    _fill_sheet(ws, week_days, kw_number, kw_year, report_nr)

    wb.save(output_path)
    print(f"✅  Saved: {output_path}  ({len(week_days)} Lerntag(e), KW {kw_number}/{kw_year})")


def create_all_berichtshefte(
    lerntage: list[Lerntag],
    output_dir: str = "berichtshefte",
    template_path: str = TEMPLATE_PATH,
) -> None:
    """Export one Berichtsheft .xlsx per calendar week found in the data."""
    os.makedirs(output_dir, exist_ok=True)
    by_kw = group_by_calendar_week(lerntage)
    for nr, (year_week, _) in enumerate(sorted(by_kw.items()), start=1):
        filename = os.path.join(output_dir, f"berichtsheft_{year_week}.xlsx")
        create_berichtsheft(lerntage, year_week, filename, report_nr=nr,
                            template_path=template_path)


# ─────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────

def main():
    print("Connecting to database …")
    lerntage = fetch_all()
    print(f"✅  {len(lerntage)} Lerntag(e) loaded.\n")

    print("=" * 60)
    print("GROUPED BY LERNFELD")
    print("=" * 60)
    for lf_id, days in group_by_lernfeld(lerntage).items():
        lf_label = days[0].lernfeld.titel if days[0].lernfeld else "Ohne Lernfeld"
        print(f"\n▶ {lf_id} – {lf_label}  ({len(days)} Tag(e))")
        for lt in days:
            print(lt)

    print("\n" + "=" * 60)
    print("GROUPED BY CALENDAR WEEK")
    print("=" * 60)
    for kw, days in group_by_calendar_week(lerntage).items():
        print(f"\n▶ {kw}  ({len(days)} Tag(e))")
        for lt in days:
            print(lt)

    print("\n" + "=" * 60)
    print("EXPORTING BERICHTSHEFTE …")
    print("=" * 60)
    create_all_berichtshefte(lerntage, output_dir="berichtshefte")


if __name__ == "__main__":
    main()